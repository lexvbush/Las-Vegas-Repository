#!/usr/bin/env python3
"""Rebuild the manifests from the master Google Sheet export.

    python3 scripts/build_manifest.py path/to/sheet.xlsx

Always parse the .xlsx export, never Drive's text conversion -- that silently
truncates (it reported 248 rows for a 732-row sheet).
"""
import collections, csv, re, sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lvlib import ROOT, load_vocab, load_merges, norm_kw

MAIN_TAB = 'NEW -Image_Source_Reference'
KW_TAB   = 'In Progress-Keyword Reference t'
COLS = ['archive', 'catalog_id', 'title_caption', 'source_url', 'topic',
        'script_pages', 'downloaded', 'in_lightroom', 'metadata_attached']
PLACE = {"Nevada", "California", "Utah", "Other States", ""}
DOC_CATEGORY = "Maps, Newspapers & Documents"
DOC_WORDS = ("newspaper", "map of", "map,", " map ", "drawing", "sketch",
             "postcard", "plans", "diagram", "advertisement", "letterhead")

# Caption words that imply a category but are not vocabulary terms themselves.
STEMS = {
    "train": "Trains & Railroads — General", "railroad": "Trains & Railroads — General",
    "depot": "Trains & Railroads — General", "locomotive": "Trains & Railroads — General",
    "mine": "Mining", "mining": "Mining", "mill": "Mining", "ore": "Mining",
    "street": "Urban & Street Scenes", "downtown": "Urban & Street Scenes",
    "store": "Commerce & Buildings", "saloon": "Commerce & Buildings",
    "hotel": "Commerce & Buildings", "bank": "Commerce & Buildings",
    "church": "Commerce & Buildings", "school": "Education",
    "ranch": "Ranching & Farming", "farm": "Ranching & Farming",
    "wagon": "Horses & Animals", "horse": "Horses & Animals",
    "mule": "Horses & Animals", "freight": "Horses & Animals",
    "flood": "Events & Disasters", "fire": "Events & Disasters",
    "sewer": "Sanitation & Utilities", "water": "Sanitation & Utilities",
    "well": "Sanitation & Utilities", "desert": "Landscape & Terrain",
    "mountain": "Landscape & Terrain", "canyon": "Landscape & Terrain",
    "portrait": "People & Crowds", "family": "People & Crowds",
    "men": "People & Crowds", "women": "People & Crowds", "crowd": "People & Crowds",
    "governor": "Government & Politics", "courthouse": "Government & Politics",
    "senator": "Government & Politics",
}


def cell(v):
    """Excel hands back numeric IDs as floats -- 17638.0 is not a catalog ID."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def key(s):
    return re.sub(r"[^a-z0-9]+", "", (s or "").lower())


def build_lookup(vocab, merges):
    """term -> category, including merge-map aliases so 'train' finds 'Trains'."""
    out = {k: cat for k, (_, cat) in vocab.items() if cat not in PLACE}
    for old, new in merges.items():
        hit = vocab.get(norm_kw(new))
        if hit and hit[1] not in PLACE:
            out.setdefault(norm_kw(old), hit[1])
    for k, v in STEMS.items():
        out.setdefault(k, v)
    return sorted(out.items(), key=lambda x: -len(x[0])), out


def main(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    rows = []
    for r in list(wb[MAIN_TAB].iter_rows(values_only=True))[1:]:
        d = {COLS[i]: cell(r[i]) for i in range(len(COLS))}
        if not (d['archive'] or d['catalog_id'] or d['title_caption']):
            continue
        if d['in_lightroom'].lower() in ('yes', 'true'):
            d['in_lightroom'] = 'Yes'
        rows.append(d)

    # keywords live on a second tab, joined by catalog id
    kw = {}
    for r in wb[KW_TAB].iter_rows(values_only=True):
        c = [cell(v) for v in r]
        if len(c) < 8 or not c[2] or c[2].lower() in ('image name / lightroom title', 'archive'):
            continue
        terms = []
        for src, sep in ((c[7], ';'), (c[4], ',')):
            for t in src.split(sep):
                t = t.strip()
                if t and t not in terms:
                    terms.append(t)
        if terms:
            kw.setdefault(key(c[2]), "; ".join(terms))

    vocab, merges = load_vocab(), load_merges()
    ordered, lookup = build_lookup(vocab, merges)

    def categorize(d):
        blob = f" {norm_kw(d['title_caption'])} {norm_kw(d['topic'])} "
        if d['archive'].lower() == 'newspaper' or any(w in blob for w in DOC_WORDS):
            return DOC_CATEGORY
        if d['keywords']:
            cats = [lookup[norm_kw(k)] for k in d['keywords'].split(';') if norm_kw(k) in lookup]
            if cats:
                return collections.Counter(cats).most_common(1)[0][0]
        found = [c for t, c in ordered if len(t) > 3 and re.search(rf"\b{re.escape(t)}", blob)]
        if found:
            return collections.Counter(found).most_common(1)[0][0]
        return "Uncategorized"

    noid = 0
    for d in rows:
        d['keywords'] = kw.get(key(d['catalog_id']), "")
        d['category'] = categorize(d)
        m = re.search(r"(\d+)", d['script_pages'])
        d['page_sort'] = int(m.group(1)) if m else 9999
        if d['catalog_id']:
            d['image_id'] = d['catalog_id']
        else:
            noid += 1
            d['image_id'] = f"NEEDS-ID-{noid:02d}"

    # two catalog ids are reused in the sheet; keep the primary field unique
    seen = collections.Counter()
    for d in rows:
        seen[d['image_id']] += 1
        if seen[d['image_id']] > 1:
            d['image_id'] = f"{d['image_id']}-{seen[d['image_id']]}"

    rows.sort(key=lambda d: (d['page_sort'], d['category'], d['archive']))
    for i, d in enumerate(rows, 1):
        d['sequence'] = i

    out = ['sequence', 'image_id', 'category', 'script_pages', 'page_sort', 'archive',
           'catalog_id', 'title_caption', 'source_url', 'keywords',
           'downloaded', 'in_lightroom', 'metadata_attached']
    for name, fields in (("editor_manifest.csv", out), ("master_manifest.csv", COLS + ['keywords'])):
        with open(ROOT / "manifest" / name, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
            w.writeheader()
            w.writerows(rows)

    cats = collections.Counter(d['category'] for d in rows)
    done = len(rows) - cats['Uncategorized']
    print(f"{len(rows)} rows | categorized {done} ({100*done//len(rows)}%) | "
          f"no catalog id: {noid} | need metadata: "
          f"{sum(1 for d in rows if d['metadata_attached'].lower() != 'yes')}")
    return rows


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "sheet.xlsx")
